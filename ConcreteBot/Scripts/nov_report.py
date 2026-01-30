#!/usr/bin/env python3
"""Generate a November concrete mix report from LineItems tabs."""

import argparse
import csv
import glob
import os
import sys
from collections import defaultdict

DEFAULT_INPUT_DIR = (
    "/Users/devinhayward/Library/CloudStorage/OneDrive-Personal/"
    "01-Active ToddGlen Projects/01-Park Properties ToddGlen/"
    "600 Lolita Gardens/Estimating/Costs to Complete/"
    "Lolita_CostToComplete_Jan_2026/Concrete Tickets/"
    "15. Nov 2025/Nov_All_Data"
)

REQUIRED_COLUMNS = [
    "Item Type",
    "Item Description",
    "Qty Value",
    "Qty Unit",
    "Unit Rate",
    "Cost",
    "Location",
    "Level",
    "Ticket No.",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a November concrete mix report from Excel LineItems tabs."
    )
    parser.add_argument(
        "--input",
        default=DEFAULT_INPUT_DIR,
        help="Directory containing Nov*.xlsx files",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output CSV path (default: <input>/Nov_Report.csv)",
    )
    return parser.parse_args()


def parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return float(stripped.replace(",", ""))
        except ValueError:
            return None
    return None


def normalize_text(value, fallback="Unknown"):
    if value is None:
        return fallback
    if isinstance(value, str):
        trimmed = value.strip()
        return trimmed if trimmed else fallback
    return str(value)


def normalize_level(value):
    if value is None:
        return "Unknown"
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, str):
        trimmed = value.strip()
        return trimmed if trimmed else "Unknown"
    return str(value)


def format_number(value):
    if value is None:
        return ""
    return f"{value:.2f}"

def level_sort_key(value):
    numeric = parse_number(value)
    if numeric is not None:
        return (0, numeric, str(value))
    return (1, str(value))


def load_line_items(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            "Missing dependency: openpyxl. Install with 'pip install openpyxl'.",
            file=sys.stderr,
        )
        sys.exit(1)

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "LineItems" not in workbook.sheetnames:
        print(f"Skipping {os.path.basename(path)}: no LineItems sheet", file=sys.stderr)
        return []
    sheet = workbook["LineItems"]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        print(f"Skipping {os.path.basename(path)}: empty LineItems", file=sys.stderr)
        return []

    header_map = {}
    for idx, name in enumerate(header):
        if name is None:
            continue
        key = str(name).strip()
        if key and key not in header_map:
            header_map[key] = idx

    missing = [col for col in REQUIRED_COLUMNS if col not in header_map]
    if missing:
        print(
            f"Skipping {os.path.basename(path)}: missing columns {missing}",
            file=sys.stderr,
        )
        return []

    items = []
    for row in rows:
        if row is None:
            continue
        items.append((row, header_map))
    return items


def build_report(input_dir, output_path):
    xlsx_paths = sorted(glob.glob(os.path.join(input_dir, "*.xlsx")))
    if not xlsx_paths:
        print(f"No .xlsx files found in {input_dir}", file=sys.stderr)
        sys.exit(1)

    grouped_main = defaultdict(lambda: {
        "total_qty": 0.0,
        "total_cost": 0.0,
        "ticket_set": set(),
        "cost_count": 0,
    })
    grouped_additional = defaultdict(lambda: {
        "total_qty": 0.0,
        "total_cost": 0.0,
        "ticket_set": set(),
        "cost_count": 0,
    })

    included_main = 0
    included_additional = 0
    skipped_desc = 0
    skipped_qty = 0
    processed_files = 0

    for path in xlsx_paths:
        items = load_line_items(path)
        if not items:
            continue
        processed_files += 1
        for row, header_map in items:
            def get_value(name):
                idx = header_map[name]
                return row[idx] if idx < len(row) else None

            item_type = normalize_text(get_value("Item Type"), fallback="").strip()
            if not item_type:
                continue

            description = normalize_text(get_value("Item Description"), fallback="").strip()
            if not description:
                skipped_desc += 1
                continue

            qty_value = parse_number(get_value("Qty Value"))
            if qty_value is None:
                skipped_qty += 1
                continue

            qty_unit = normalize_text(get_value("Qty Unit"), fallback="Unknown")
            location = normalize_text(get_value("Location"), fallback="Unknown")
            level = normalize_level(get_value("Level"))
            ticket_no = normalize_text(get_value("Ticket No."), fallback="")

            unit_rate = parse_number(get_value("Unit Rate"))
            cost = parse_number(get_value("Cost"))
            if cost is not None and cost > 0:
                computed_cost = cost
            elif unit_rate is not None:
                computed_cost = qty_value * unit_rate
            else:
                computed_cost = None

            if item_type == "Mix Customer":
                key = (location, level, description, qty_unit)
                agg = grouped_main[key]
                included_main += 1
            else:
                key = (location, level, item_type, description, qty_unit)
                agg = grouped_additional[key]
                included_additional += 1
            agg["total_qty"] += qty_value
            if computed_cost is not None:
                agg["total_cost"] += computed_cost
                agg["cost_count"] += 1
            if ticket_no:
                agg["ticket_set"].add(ticket_no)

    with open(output_path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        header = [
            "Level",
            "Location",
            "Item Type",
            "Mix Description",
            "Ticket Count",
            "Total Qty",
            "Qty Unit",
            "Unit Rate",
            "Total Cost",
        ]

        writer.writerow(["Main Mixes"])
        writer.writerow(header)
        for (location, level, description, qty_unit), agg in sorted(
            grouped_main.items(),
            key=lambda item: (
                level_sort_key(item[0][1]),
                item[0][0],
                item[0][2],
                item[0][3],
            ),
        ):
            total_qty = agg["total_qty"]
            total_cost = agg["total_cost"] if agg["cost_count"] > 0 else None
            avg_unit_rate = None
            if total_cost is not None and total_qty > 0:
                avg_unit_rate = total_cost / total_qty
            writer.writerow([
                level,
                location,
                "Mix Customer",
                description,
                len(agg["ticket_set"]),
                format_number(total_qty),
                qty_unit,
                format_number(avg_unit_rate),
                format_number(total_cost),
            ])

        writer.writerow([])
        writer.writerow(["Additional Mixes"])
        writer.writerow(header)
        for (location, level, item_type, description, qty_unit), agg in sorted(
            grouped_additional.items(),
            key=lambda item: (
                level_sort_key(item[0][1]),
                item[0][0],
                item[0][2],
                item[0][3],
                item[0][4],
            ),
        ):
            total_qty = agg["total_qty"]
            total_cost = agg["total_cost"] if agg["cost_count"] > 0 else None
            avg_unit_rate = None
            if total_cost is not None and total_qty > 0:
                avg_unit_rate = total_cost / total_qty
            writer.writerow([
                level,
                location,
                item_type,
                description,
                len(agg["ticket_set"]),
                format_number(total_qty),
                qty_unit,
                format_number(avg_unit_rate),
                format_number(total_cost),
            ])

    print(f"Processed files: {processed_files}")
    print(f"Included rows (Main Mixes): {included_main}")
    print(f"Included rows (Additional Mixes): {included_additional}")
    print(f"Skipped rows (blank Item Description): {skipped_desc}")
    print(f"Skipped rows (non-numeric Qty Value): {skipped_qty}")
    print(f"Report written to: {output_path}")


def main():
    args = parse_args()
    input_dir = os.path.expanduser(args.input)
    output_path = args.output
    if output_path is None:
        output_path = os.path.join(input_dir, "Nov_Report.csv")
    output_path = os.path.expanduser(output_path)

    if not os.path.isdir(input_dir):
        print(f"Input directory not found: {input_dir}", file=sys.stderr)
        sys.exit(1)

    build_report(input_dir, output_path)


if __name__ == "__main__":
    main()
